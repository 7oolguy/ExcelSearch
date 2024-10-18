""" import openpyxl
import openpyxl.utils
import openpyxl.utils.exceptions
import openpyxl.worksheet
import openpyxl.worksheet.worksheet


class ExcelManager:
    def __init__(self, file_path=None, sheet_name=None):
        self._file_path = file_path
        self._sheet_name = sheet_name
        self._workbook = None
        self._sheet = None

    def __str__(self) -> str:
        return self.file_path

    def __eq__(self, value: object) -> bool:
        return self.file_path == value or self.sheet_name == value

    @property
    def file_path(self):
        return self._file_path

    @file_path.setter
    def file_path(self, file_path):
        if isinstance(file_path, str):
            self._file_path = file_path
        else:
            raise ValueError("File Path needs to be a string.")

    @property
    def sheet_name(self):
        return self._sheet_name

    @sheet_name.setter
    def sheet_name(self, sheet_name):
        if isinstance(sheet_name, str):
            self._sheet_name = sheet_name
        else:
            raise ValueError("Sheet Name needs to be a string.")

    @property
    def workbook(self):
        return self._workbook

    @workbook.setter
    def workbook(self, workbook):
        if isinstance(workbook, openpyxl.Workbook):
            self._workbook = workbook
        else:
            raise ValueError("The value must be an instance of openpyxl Workbook.")

    @property
    def sheet(self):
        return self._sheet

    @sheet.setter
    def sheet(self, sheet):
        if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            self._sheet = sheet
        else:
            raise ValueError("The value must be an instance of openpyxl Worksheet")

    def load(self, file_path: str, sheet_name: str = None) -> None:
        try:
            self.file_path = file_path
            self.workbook = openpyxl.load_workbook(file_path)
            if sheet_name:
                if sheet_name in self.workbook.sheetnames:
                    self.sheet_name = sheet_name
                else:
                    raise ValueError("Sheet does not exists in the workbook.")
            else:
                self.sheet_name = self.workbook.sheetnames[0]
            self.sheet = self.workbook[self.sheet_name]
        except FileNotFoundError:
            raise FileNotFoundError(f"The file '{file_path}' was not found.")
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError("The file provided is not a valid Excel file.")

    def save(self):
        self.workbook.save(self.file_path)

    def get_sheet_names(self) -> list:
        if self.workbook:
            return self.workbook.sheetnames
        else:
            raise ValueError("No workbook is currently loaded.")

    def get_table_names(self) -> list:
        if self.sheet:
            return list(self.sheet.tables.keys())
        else:
            raise ValueError("No sheet is currently loaded.")

    def change_cell_value(self, row: int, col: int, value: str):
        if not self.sheet:
            raise ValueError("No sheet is currently loaded")

        if row and col:
            self.sheet.cell(row=row, column=col).value = value
        else:
            raise ValueError("You must provide the row number and the column number.")

        self.save()

    def get_header(self, table_name: str = None) -> list:
        if not self.sheet:
            return []

        if table_name:
            if table_name not in self.sheet.tables:
                raise ValueError(f"Table '{table_name}' does not exists in the sheet.")

            table = self.sheet.tables[table_name]
            table_range = self.sheet[table.ref]
            header = [cell.value for cell in next(iter(table_range))]
        else:
            header = [cell for cell in next(self.sheet.iter_rows(values_only=True))]

        return header

    def get_rows(self, table_name: str = None) -> list:
        if not self.sheet:
            print("No sheet found.")
            return [[]]

        rows = []

        if table_name:
            if table_name not in self.sheet.tables:
                raise ValueError(f"Table '{table_name}' does not exist in the sheet.")

            table = self.sheet.tables[table_name]
            table_range = self.sheet[table.ref]

            for i, row in enumerate(table_range):
                row_values = [cell.value for cell in row if cell.value is not None]
                print(f"Row {i}: {row_values}")  # Debugging output
                rows.append(row_values)
        else:
            for i, row in enumerate(self.sheet.iter_rows(values_only=True), 1):
                row_values = [cell for cell in row if cell is not None]
                print(f"Row {i}: {row_values}")  # Debugging output
                rows.append(row_values)

        return rows

    def get_value_row(self, value: str, table_name: str = None) -> list[list]:
        rows = self.get_rows(table_name=table_name)  # Fetch rows without headers
        for row in rows:
            print(f"Fetched Rows: {row}")  # Debugging output
        value_low = value.lower()
        matched_rows = []

        if rows:
            for row in rows:
                if any(value_low in str(item).lower() for item in row if item is not None):
                    matched_rows.append(row)
                    print(f"Matched Row: {row}")  # Debugging output

        return matched_rows
 """
import pandas as pd


class ExcelManager:
    def __init__(self, file_path: str = None, sheet_name: str = None) -> None:
        self._file_path = file_path
        self._sheet_name = sheet_name
        self._dataframe = None

    def load(self, file_path: str, sheet_name: str = None) -> None:
        self._file_path = file_path
        if sheet_name:
            self._sheet_name = sheet_name
            self._dataframe = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            self._dataframe = pd.read_excel(file_path)
            self._sheet_name = self._dataframe.columns.name

    def save(self, file_path: str = None):
        save_path = file_path if file_path else self._file_path
        self._dataframe.to_excel(save_path, index=False)

    def get_sheet_names(self) -> list:
        return pd.ExcelFile(self._file_path).sheet_names

    def get_table_names(self) -> list:
        # Pandas does not have this funcitonality
        return

    def change_cell_value(self, row: int, col: int, value: str):
        if self._dataframe is not None:
            self._dataframe.iloc[row - 1, col - 1] = value()
        else:
            raise ValueError("No sheet is currently loaded")

    def get_header(self) -> list:
        if self._dataframe is not None:
            return self._dataframe.columns.tolist()
        else:
            return []

    def get_rows(self) -> list:
        if self._dataframe is not None:
            return self._dataframe.values.tolist()
        return []

    def get_value_row(self, value: str, table_name: str = None) -> list[list]:
        # Ensure the dataframe is loaded
        if self._dataframe is None:
            raise ValueError("No sheet is loaded.")

        # Perform a case-insensitive search across the entire DataFrame
        value_lower = value.lower()

        # Apply a function row-by-row to match cells that either match exactly or start with the search term
        mask = self._dataframe.apply(
            lambda row: row.astype(str).apply(
                lambda x: (x.lower() == value_lower or x.lower().startswith(value_lower)) if pd.notna(x) else False
            ),
            axis=1
        )

        # Get rows where any cell in the row matches the condition
        matching_rows = self._dataframe[mask.any(axis=1)]

        # Convert the result to a list of lists
        return matching_rows.values.tolist()
